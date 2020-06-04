from openpyxl import load_workbook
import json


wb = load_workbook(filename = 'TestingOrder.xlsx', data_only=True)
ws = wb["Testing Order"]
colNames = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
numfactors = input("Input the number of factors that were tested: ")
with open("key.json", "r") as f:
    key = json.load(f)

dataCol = input("Please enter the column where you have stored the data: ")
dataCol.upper()
data = []
for row in ws[dataCol]:
    data.append(row.value)

runData = []
headers = []
for header in ws['1'][:int(numfactors)]:
    headers.append(header.value)
runs = len(ws['A']) # Subtract one to include the header
for i in range(2, runs + 1):
    run = {}
    states = ws[str(i)]
    for j in range(int(numfactors)):
        run[headers[j]] = key[headers[j]][str(states[j].value)]
    run["time"] = data[i-1]
    runData.append(run)
print('\n\n--------MAIN EFFECTS---------\n\n')


for i in runData[0].keys():
    if i == 'time':
        break
    highSum = 0
    lowSum = 0
    highCnt = 0
    lowCnt = 0
    
    for j in runData:
        if j[i] == '1':
            highSum += j['time']
            highCnt += 1
        else:
            lowSum += j['time']
            lowCnt += 1
    lowAvg = lowSum/lowCnt
    highAvg = highSum/highCnt
    print(f"Average {data[0]} with {i} on 'low': {lowAvg} -- Average {data[0]} with {i} on 'high': {highAvg}")
    print(f"Going from 'high' to 'low' {i} {'increases' if (highAvg - lowAvg) < 0 else 'decreases'} {data[0]} by {abs(round(highAvg - lowAvg,3))} +/- 0.002 seconds")
    print('\n-----------------------\n')
print('\n\n----------INTERACTIONS-------------\n\n')

for i in runData[0].keys():
    if i == 'time':
        break
    
    for j in runData[0].keys():
        if j == 'time' or i == j:
            break
        hhsum = hhcnt = hlsum = hlcnt = lhsum = lhcnt = llsum = llcnt =  0
        for k in runData:
            if k[i] == '1' and k[j] == '1':
                hhsum += k['time']
                hhcnt += 1
            elif k[i] == '1' and k[j] == '-1':
                hlsum += k['time']
                hlcnt += 1
            elif k[i] == '-1' and k[j] == '1':
                lhsum += k['time']
                lhcnt += 1
            else:
                llsum += k['time']
                llcnt += 1
        hhavg = round(hhsum / hhcnt, 3)
        hlavg = round(hlsum / hhcnt, 3)
        lhavg = round(lhsum / lhcnt, 3)
        llavg = round(llsum / llcnt, 3)
        print(f"Average {data[0]} with {i} on the high setting and {j} on the high setting = {hhavg}")
        print(f"Average {data[0]} with {i} on the high setting and {j} on the low setting = {hlavg}")
        print(f"{j} effects {i}'s effect on {data[0]} by making it {abs(round(hhavg - hlavg, 3))} +/- 0.002 seconds {'slower' if (hhavg - hlavg) < 0 else 'faster'}, going from 'high' {j} to 'low' {j}")
        print('\n-----------------------\n')
        print(f"Average {data[0]} with {i} on the low setting and {j} on the high setting = {lhavg}")
        print(f"Average {data[0]} with {i} on the low setting and {j} on the low setting = {llavg}")
        print(f"{j} effects {i}'s effect on {data[0]} by making it {abs(round(lhavg - llavg, 3))} +/- 0.002 seconds {'slower' if (lhavg - llavg) < 0 else 'faster'}, going from 'high' {j} to 'low' {j}")
        print('\n-----------------------\n')





