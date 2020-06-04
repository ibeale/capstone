from openpyxl import load_workbook
import json
import matplotlib.pyplot as plt

def main():
    results = open("results.txt", "w")
    wb = load_workbook(filename = 'TestingOrder.xlsx', data_only=True)
    ws = wb["Testing Order"]
    colNames = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    numfactors = input("Input the number of factors that were tested: ")
    with open("key.json", "r") as f:
        key = json.load(f)

    dataCol = input("Please enter the column where you have stored the data: ")
    dataCol.upper()
    data = []
    for row in ws[dataCol]:
        data.append(row.value)

    runData = []
    headers = []
    for header in ws['1'][:int(numfactors)]:
        headers.append(header.value)
    runs = len(ws['A']) # Subtract one to include the header
    for i in range(2, runs + 1):
        run = {}
        states = ws[str(i)]
        for j in range(int(numfactors)):
            run[headers[j]] = key[headers[j]][str(states[j].value)]
        run["time"] = data[i-1]
        runData.append(run)
    print('\n\n--------MAIN EFFECTS---------\n\n')
    results.write('\n\n--------MAIN EFFECTS---------\n\n')

    min_avg = 1000000
    max_avg = 0
    x_vals = [-1,1]
    for idx, i in enumerate(runData[0].keys()):
        if i == 'time':
            break
        highSum = 0
        lowSum = 0
        highCnt = 0
        lowCnt = 0
        
        for j in runData:
            if j[i] == '1':
                highSum += j['time']
                highCnt += 1
            else:
                lowSum += j['time']
                lowCnt += 1
        lowAvg = lowSum/lowCnt
        highAvg = highSum/highCnt
        if(highAvg > max_avg or lowAvg > max_avg):
            max_avg = highAvg if (highAvg > lowAvg) else lowAvg
        if(highAvg < max_avg or lowAvg < max_avg):
            min_avg = highAvg if (highAvg < lowAvg) else lowAvg
        y_vals = [lowAvg,highAvg]
        plt.plot(x_vals, y_vals, label = i)
        print(f"Average {data[0]} with {i} on 'low': {round(lowAvg,3)} -- Average {data[0]} with {i} on 'high': {round(highAvg,3)}")
        print(f"Going from 'high' to 'low' {i} {'increases' if (highAvg - lowAvg) < 0 else 'decreases'} {data[0]} by {abs(round(highAvg - lowAvg,3))} +/- 0.002 seconds")
        print('\n-----------------------\n')
        results.write(f"Average {data[0]} with {i} on 'low': {round(lowAvg,3)} -- Average {data[0]} with {i} on 'high': {round(highAvg,3)}\n")
        results.write(f"Going from 'high' to 'low' {i} {'increases' if (highAvg - lowAvg) < 0 else 'decreases'} {data[0]} by {abs(round(highAvg - lowAvg,3))} +/- 0.002 seconds\n")
        results.write('\n-----------------------\n')
    plt.title('Main Effects Plots')
    plt.legend()
    print("Close the graph to continue")
    plt.show()
    print('\n\n----------INTERACTIONS-------------\n\n')
    results.write('\n\n----------INTERACTIONS-------------\n\n')

    for i in runData[0].keys():
        if i == 'time':
            break
        
        for j in runData[0].keys():
            if j == 'time' or i == j:
                break
            hhsum = hhcnt = hlsum = hlcnt = lhsum = lhcnt = llsum = llcnt =  0
            for k in runData:
                if k[i] == '1' and k[j] == '1':
                    hhsum += k['time']
                    hhcnt += 1
                elif k[i] == '1' and k[j] == '-1':
                    hlsum += k['time']
                    hlcnt += 1
                elif k[i] == '-1' and k[j] == '1':
                    lhsum += k['time']
                    lhcnt += 1
                else:
                    llsum += k['time']
                    llcnt += 1
            hhavg = round(hhsum / hhcnt, 3)
            hlavg = round(hlsum / hhcnt, 3)
            lhavg = round(lhsum / lhcnt, 3)
            llavg = round(llsum / llcnt, 3)
            plt.plot([-1,1], [hlavg, hhavg], label = f"High {i}")
            plt.plot([-1,1], [llavg, lhavg], label = f"Low {i}")
            plt.title(f'Interaction between {i} and {j}')
            plt.legend()
            print(f"Average {data[0]} with {i} on the high setting and {j} on the high setting = {hhavg}")
            print(f"Average {data[0]} with {i} on the high setting and {j} on the low setting = {hlavg}")
            print(f"{j} effects {i}'s effect on {data[0]} by making it {abs(round(hhavg - hlavg, 3))} +/- 0.002 seconds {'slower' if (hhavg - hlavg) < 0 else 'faster'}, going from 'high' {j} to 'low' {j}")
            print('\n-----------------------\n')
            print(f"Average {data[0]} with {i} on the low setting and {j} on the high setting = {lhavg}")
            print(f"Average {data[0]} with {i} on the low setting and {j} on the low setting = {llavg}")
            print(f"{j} effects {i}'s effect on {data[0]} by making it {abs(round(lhavg - llavg, 3))} +/- 0.002 seconds {'slower' if (lhavg - llavg) < 0 else 'faster'}, going from 'high' {j} to 'low' {j}")
            print('\n-----------------------\n')
            results.write(f"Average {data[0]} with {i} on the high setting and {j} on the high setting = {hhavg}\n")
            results.write(f"Average {data[0]} with {i} on the high setting and {j} on the low setting = {hlavg}\n")
            results.write(f"{j} effects {i}'s effect on {data[0]} by making it {abs(round(hhavg - hlavg, 3))} +/- 0.002 seconds {'slower' if (hhavg - hlavg) < 0 else 'faster'}, going from 'high' {j} to 'low' {j}\n")
            results.write('\n-----------------------\n')
            results.write(f"Average {data[0]} with {i} on the low setting and {j} on the high setting = {lhavg}\n")
            results.write(f"Average {data[0]} with {i} on the low setting and {j} on the low setting = {llavg}\n")
            results.write(f"{j} effects {i}'s effect on {data[0]} by making it {abs(round(lhavg - llavg, 3))} +/- 0.002 seconds {'slower' if (lhavg - llavg) < 0 else 'faster'}, going from 'high' {j} to 'low' {j}\n")
            results.write('\n-----------------------\n')
            print("Close the graph to continue. Note: intersecting lines indicate an interaction. Parallel lines indicate no interaction")
            plt.show()
    results.close()

if __name__ == "__main__":
    main()



